"""
regenerar_informe.py
=====================
Regenera el informe de resultados a partir de los JSON de detalle
(detalle_iXX.json) que ya contienen TODOS los costes del validador,
incluidos ElectiveUnscheduledPatients y PatientDelay.

No re-ejecuta el pipeline ni el validador: solo reconstruye las tablas.

Genera DOS ficheros a prueba de Excel:
  · informe.xlsx  → formato nativo Excel (sin corrupción de decimales/fechas)
  · informe.csv   → CSV UTF-8 con separador coma y punto decimal

Por defecto procesa la carpeta del run rápido. Cambia --carpeta para otra.

Uso:
  python regenerar_informe.py
  python regenerar_informe.py --carpeta resultados_maximo_benchmark
"""

import argparse
import csv
import json
from pathlib import Path

BASE_DIR = Path(__file__).parent.resolve()

# Componentes de coste del IHTP_Validator (nombres exactos), de mayor a menor
# relevancia típica.
COMPONENTES_COSTE = [
    "ElectiveUnscheduledPatients",
    "PatientDelay",
    "RoomSkillLevel",
    "ContinuityOfCare",
    "OpenOperatingTheater",
    "RoomAgeMix",
    "ExcessiveNurseWorkload",
    "SurgeonTransfer",
]

COLUMNAS = [
    "instancia", "estado", "total_violations", "total_cost",
    "pacientes_f1", "pacientes_f2", "pacientes_eliminados_poda",
    "t_fase1", "t_fase2", "t_fase3", "t_total",
] + [f"cost_{c}" for c in COMPONENTES_COSTE] + [
    "score_interno", "threads", "mip_gap", "timestamp",
]


def _g(d, *keys, default=""):
    """Busca una clave en el dict y, si no, dentro de 'metricas'."""
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
        m = d.get("metricas", {})
        if k in m and m[k] is not None:
            return m[k]
    return default


def fila_desde_detalle(d):
    t = d.get("tiempos", {})
    fila = {
        "instancia":                 d.get("instancia", ""),
        "estado":                    d.get("estado", ""),
        "total_violations":          _g(d, "total_violations"),
        "total_cost":                _g(d, "total_cost"),
        "pacientes_f1":              _g(d, "pacientes_f1"),
        "pacientes_f2":              _g(d, "pacientes_f2"),
        "pacientes_eliminados_poda": _g(d, "pacientes_eliminados_poda", default=0),
        "t_fase1":                   t.get("fase1", ""),
        "t_fase2":                   t.get("fase2", ""),
        "t_fase3":                   t.get("fase3", ""),
        "t_total":                   t.get("total", ""),
        "score_interno":             _g(d, "score_interno"),
        "threads":                   d.get("threads", ""),
        "mip_gap":                   d.get("mip_gap", ""),
        "timestamp":                 d.get("timestamp", ""),
    }
    for c in COMPONENTES_COSTE:
        fila[f"cost_{c}"] = _g(d, f"cost_{c}", default=0)
    return fila


def main():
    parser = argparse.ArgumentParser(description="Regenera informe desde JSON de detalle.")
    parser.add_argument("--carpeta", default="resultados_rapido_benchmark",
                        help="Carpeta con los detalle_iXX.json")
    args = parser.parse_args()

    carpeta = BASE_DIR / args.carpeta
    if not carpeta.exists():
        print(f"[ERROR] No existe la carpeta: {carpeta}")
        return

    detalles = sorted(carpeta.glob("detalle_i*.json"))
    if not detalles:
        print(f"[ERROR] No hay ficheros detalle_iXX.json en {carpeta}")
        return

    filas = []
    for p in detalles:
        try:
            d = json.load(open(p, encoding="utf-8"))
            filas.append(fila_desde_detalle(d))
        except Exception as e:
            print(f"[AVISO] No se pudo leer {p.name}: {e}")

    filas.sort(key=lambda f: f["instancia"])

    # ── CSV UTF-8 (separador coma, punto decimal) ─────────────────────
    ruta_csv = carpeta / "informe.csv"
    with open(ruta_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNAS)
        w.writeheader()
        for fila in filas:
            w.writerow(fila)

    # ── XLSX nativo (Excel no corrompe nada) ──────────────────────────
    ruta_xlsx = carpeta / "informe.xlsx"
    xlsx_ok = False
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        cabecera_fill = PatternFill("solid", fgColor="4472C4")
        cabecera_font = Font(bold=True, color="FFFFFF")

        ws.append(COLUMNAS)
        for c in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = cabecera_fill
            cell.font = cabecera_font
            cell.alignment = Alignment(horizontal="center")

        for fila in filas:
            ws.append([fila.get(col, "") for col in COLUMNAS])

        # Anchos de columna razonables
        for c, col in enumerate(COLUMNAS, 1):
            ws.column_dimensions[get_column_letter(c)].width = max(12, len(col) + 2)

        ws.freeze_panes = "B2"
        wb.save(ruta_xlsx)
        xlsx_ok = True
    except ImportError:
        print("[AVISO] openpyxl no instalado: omito el .xlsx (instala con: pip install openpyxl)")

    # ── Resumen por pantalla ──────────────────────────────────────────
    print("\n" + "=" * 96)
    print("  INFORME REGENERADO  (costes reales del validador)")
    print("=" * 96)
    hdr = (f"{'Inst':>5} {'Estado':>8} {'Viol':>5} {'Coste':>8} "
           f"{'Unsched':>8} {'Delay':>7} {'Skill':>6} {'Contin':>7} "
           f"{'OpenOT':>7} {'Poda':>5} {'t(s)':>7}")
    print(hdr)
    print("-" * 96)

    sum_unsched = sum_delay = sum_total = 0
    for fila in filas:
        unsched = fila.get("cost_ElectiveUnscheduledPatients", 0) or 0
        delay   = fila.get("cost_PatientDelay", 0) or 0
        skill   = fila.get("cost_RoomSkillLevel", 0) or 0
        contin  = fila.get("cost_ContinuityOfCare", 0) or 0
        openot  = fila.get("cost_OpenOperatingTheater", 0) or 0
        total   = fila.get("total_cost", 0) or 0
        poda    = fila.get("pacientes_eliminados_poda", 0) or 0
        t       = fila.get("t_total", "")
        try:    sum_total += int(total)
        except: pass
        try:    sum_unsched += int(unsched)
        except: pass
        try:    sum_delay += int(delay)
        except: pass
        print(f"{fila['instancia']:>5} {str(fila['estado']):>8} "
              f"{str(fila['total_violations']):>5} {str(total):>8} "
              f"{str(unsched):>8} {str(delay):>7} {str(skill):>6} "
              f"{str(contin):>7} {str(openot):>7} {str(poda):>5} {str(t):>7}")

    print("-" * 96)
    print(f"  Coste total acumulado            : {sum_total}")
    if sum_total:
        print(f"  De ello, ElectiveUnscheduled     : {sum_unsched}  ({sum_unsched/sum_total*100:.0f}% del total)")
        print(f"  De ello, PatientDelay            : {sum_delay}  ({sum_delay/sum_total*100:.0f}% del total)")
    print("=" * 96)
    print(f"\n  CSV  : {ruta_csv}")
    if xlsx_ok:
        print(f"  XLSX : {ruta_xlsx}   (ábrelo en Excel: ya no corrompe decimales ni fechas)")


if __name__ == "__main__":
    main()
